"""Seed the SQLite database from the Exercises spreadsheet.

Reads the 'Sessions' sheet for routines + exercises (and the Session-1 S1
weight as a starting weight) and the 'Personal Best' sheet for PB weight/reps.
Idempotent-ish: only seeds when the routine table is empty.
"""
import os

import openpyxl

from db import get_db, init_db, is_seeded

XLSX_PATH = os.environ.get(
    "GYM_XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "..", "Exercises_2.xlsx"),
)


def _num(value):
    """Return a float for numeric cells, else None (handles 'N/A', blanks)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _read_personal_bests(wb):
    """Map exercise name -> (pb_weight, pb_reps) from the Personal Best sheet."""
    pbs = {}
    if "Personal Best" not in wb.sheetnames:
        return pbs
    ws = wb["Personal Best"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not name:
            continue
        weight = _num(row[2]) if len(row) > 2 else None
        reps = row[3] if len(row) > 3 else None
        reps = int(reps) if isinstance(reps, (int, float)) else None
        pbs[str(name).strip()] = (weight, reps)
    return pbs


def seed():
    init_db()
    if is_seeded():
        print("Database already seeded; skipping.")
        return

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Sessions"]
    pbs = _read_personal_bests(wb)

    conn = get_db()
    cur = conn.cursor()

    current_routine_id = None
    routine_pos = 0
    ex_pos = 0

    # Data rows start at row 3 (1-based): row 1 = session headers, row 2 = column headers.
    for row in ws.iter_rows(min_row=3, values_only=True):
        routine_name = (row[0] or "").strip() if row[0] else ""
        exercise_name = (row[1] or "").strip() if row[1] else ""
        notes = row[2].strip() if (len(row) > 2 and isinstance(row[2], str)) else None
        start_weight = _num(row[4]) if len(row) > 4 else None  # Session 1 -> S1 Wt

        if routine_name:
            routine_pos += 1
            ex_pos = 0
            cur.execute(
                "INSERT INTO routine (name, position) VALUES (?, ?)",
                (routine_name, routine_pos),
            )
            current_routine_id = cur.lastrowid

        if not exercise_name or current_routine_id is None:
            continue

        ex_pos += 1
        pb_weight, pb_reps = pbs.get(exercise_name, (None, None))
        cur.execute(
            """INSERT INTO exercise
                   (routine_id, name, notes, position, start_weight, pb_weight, pb_reps)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (current_routine_id, exercise_name, notes, ex_pos, start_weight, pb_weight, pb_reps),
        )

    conn.commit()
    n_routines = conn.execute("SELECT COUNT(*) FROM routine").fetchone()[0]
    n_exercises = conn.execute("SELECT COUNT(*) FROM exercise").fetchone()[0]
    conn.close()
    print(f"Seeded {n_routines} routines and {n_exercises} exercises from {XLSX_PATH}")


if __name__ == "__main__":
    seed()
